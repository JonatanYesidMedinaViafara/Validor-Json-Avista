import pandas as pd
from pathlib import Path
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher


class NormalizadorExcel:
    def __init__(self, carpeta_excel_origen: str, carpeta_excel_destino: str, carpeta_excel_fallos: str, umbral_similitud: float = 0.7):
        self.carpeta_excel_origen = Path(carpeta_excel_origen)
        self.carpeta_excel_destino = Path(carpeta_excel_destino)
        self.carpeta_excel_fallos = Path(carpeta_excel_fallos)
        self.logger = logging.getLogger("NormalizadorExcel")
        self.umbral_similitud = umbral_similitud  # <-- nuevo parámetro para tolerancia

    def _similitud_texto(self, a: str, b: str) -> float:
        """Calcula el porcentaje de similitud entre dos textos."""
        return SequenceMatcher(None, a.strip().lower(), b.strip().lower()).ratio()

    def _validar_bloque(self, row, columnas, nombre_unificado):
        """Valida un bloque comparando con tolerancia (similitud promedio)."""
        valores = [str(row[col]).strip() for col in columnas if col in row.index and pd.notna(row[col])]
        if not valores:
            return None, columnas  # Si no hay datos, lo marcamos como error
        comparaciones = []
        for i in range(len(valores)):
            for j in range(i + 1, len(valores)):
                similitud = self._similitud_texto(valores[i], valores[j])
                comparaciones.append(similitud)

        if not comparaciones:
            return valores[0], []  # Solo un valor → lo aceptamos

        similitud_promedio = sum(comparaciones) / len(comparaciones)

        if similitud_promedio >= self.umbral_similitud:
            return valores[0], []  # Si cumple el umbral → válido
        else:
            # Devolver columnas conflictivas
            return None, columnas

    def _obtener_ultimo_excel(self):
        archivos = list(self.carpeta_excel_origen.glob("clon_json_*_reestructurado.xlsx"))
        if not archivos:
            self.logger.error("No se encontraron archivos reestructurados.")
            return None
        ultimo = max(archivos, key=lambda f: f.stat().st_mtime)
        return ultimo

    def _validar_bloque(self, row, columnas, nombre_unificado):
        """Valida si las columnas del bloque tienen el mismo valor. Retorna el valor unificado y las celdas con error."""
        valores = [str(row[col]).strip() for col in columnas if col in row.index]
        valores_unicos = set(valores)
        if len(valores_unicos) == 1:
            return valores_unicos.pop(), []  # Correcto
        else:
            valores_base = valores[0]
            columnas_error = [col for col, val in zip(columnas, valores) if val != valores_base]
            return None, columnas_error

    def _convertir_numeros(self, df):
        """Convierte columnas numéricas de texto a números, exceptuando fechas."""
        columnas_excluir = ["cedula_fecha_nacimiento", "desprendible_nomina_vigencia"]
        for col in df.columns:
            if col not in columnas_excluir:
                try:
                    df[col] = pd.to_numeric(df[col], errors='ignore')
                except:
                    pass
        return df

    def _pintar_errores(self, ruta_excel, columnas_error):
        """Pinta en rojo las celdas con error en el Excel de fallos."""
        wb = load_workbook(ruta_excel)
        ws = wb.active
        fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        header = [cell.value for cell in ws[1]]
        col_indices = [header.index(col) + 1 for col in columnas_error if col in header]

        for row in ws.iter_rows(min_row=2):
            for idx in col_indices:
                row[idx-1].fill = fill

        wb.save(ruta_excel)

    def _renombrar_columnas(self, df):
        """Renombra columnas específicas según el requerimiento."""
        renombres = {
            "amortizacion_tasa_interes": "Tasa Interes",
            "cedula_fecha_nacimiento": "Fecha de Nacimiento",
            "desprendible_nomina_salario": "Desprendible de Nomina",
            "id_cargue_origen":"Numero Cargue"
        }
        df.rename(columns=renombres, inplace=True)
        return df

    def normalizar(self):
        # 1. Buscar el último Excel reestructurado
        archivo = self._obtener_ultimo_excel()
        if not archivo:
            return False

        self.logger.info(f"Normalizando archivo: {archivo.name}")

        # 2. Cargar datos
        df = pd.read_excel(archivo, dtype=str)

        # 3. Convertir a números donde aplique
        df = self._convertir_numeros(df)

        # --- Definir bloques de validación ---
        bloques = [
            {
                "columnas": ["desprendible_nomina_pagaduria", "amortizacion_pagaduria", "libranza_pagaduria"],
                "nombre_unificado": "Pagaduria Unificada"
            },
            {
                "columnas": ["formato_conocimiento_plazo_meses", "amortizacion_plazo_meses", "libranza_plazo"],
                "nombre_unificado": "Plazo Unificado"
            },
            {
                "columnas": [
                    "libranza_nombre_firma_electronica",
                    "seguro_de_vida_nombre_firma_electronica",
                    "solicitud_credito_nombre_firma_electronica",
                    "solicitud_fianza_nombre_firma_electronica",
                    "formato_conocimiento_nombre_firma_electronica",
                    "amortizacion_nombre_firma_electronica"
                ],
                "nombre_unificado": "Firma Electrónica Unificada"
            },
            {
                "columnas": [
                    "formato_conocimiento_cedula_firma_electronica",
                    "seguro_de_vida_cedula_firma_electronica",
                    "solicitud_fianza_cedula_firma_electronica",
                    "solicitud_credito_cedula_firma_electronica",
                    "libranza_cedula_firma_electronica",
                    "amortizacion_cedula_firma_electronica"
                ],
                "nombre_unificado": "Cédula Firma Electrónica Unificada"
            },
            {
                "columnas": [
                    "libranza_valor_prestamo",
                    "amortizacion_valor_credito",
                    "formato_conocimiento_valor_total_credito"
                ],
                "nombre_unificado": "Valor Crédito Unificado"
            },
            {
                "columnas": [
                    "libranza_valor_cuota",
                    "amortizacion_valor_cuota"
                ],
                "nombre_unificado": "Valor Cuota Unificado"
            # },
            # {
            #     "columnas": [
            #         "solicitud_credito_numero_credito",
            #         "libranza_numero_credito",
            #         "Numero credito"
            #     ],
            #     "nombre_unificado": "Número Crédito Unificado"
            },
            {
                "columnas": [
                    "solicitud_credito_solicitud",
                    "amortizacion_numero_solicitud"
                ],
                "nombre_unificado": "Número Solicitud Unificado"
            },
            {
                "columnas": [
                    "Solicitud Fianza Segundo Apellido",
                    "Solicitud Credito Segundo Apellido",
                    "Seguro De Vida Segundo Apellido",
                    "Libranza Segundo Apellido",
                    "Desprendible Nomina Segundo Apellido",
                    "Cedula Segundo Apellido",
                    "Amortizacion Segundo Apellido"
                ],
                "nombre_unificado": "Segundo Apellido Unificado"
            # },
            # {
            #     "columnas": [
            #         "Solicitud Fianza Cedula",
            #         "Solicitud Credito Cedula",
            #         "Seguro De Vida Cedula",
            #         "Libranza Cedula",
            #         "Desprendible Nomina Cedula",
            #         "Cedula Cedula",
            #         "Amortizacion Cedula",
            #         "Cedula"
            #     ],
            #     "nombre_unificado": "Cédula Unificada"
            },
            {
                "columnas": [
                    "Solicitud Fianza Primer Apellido",
                    "Solicitud Credito Primer Apellido",
                    "Seguro De Vida Primer Apellido",
                    "Libranza Primer Apellido",
                    "Desprendible Nomina Primer Apellido",
                    "Cedula Primer Apellido",
                    "Amortizacion Primer Apellido"
                ],
                "nombre_unificado": "Primer Apellido Unificado"
            },
            {
                "columnas": [
                    "Solicitud Fianza Segundo Nombre",
                    "Solicitud Credito Segundo Nombre",
                    "Seguro De Vida Segundo Nombre",
                    "Libranza Segundo Nombre",
                    "Desprendible Nomina Segundo Nombre",
                    "Cedula Segundo Nombre",
                    "Amortizacion Segundo Nombre"
                ],
                "nombre_unificado": "Segundo Nombre Unificado"
             },
             {
                "columnas": [
                    "Solicitud Fianza Primer Nombre",
                    "Solicitud Credito Primer Nombre",
                    "Seguro De Vida Primer Nombre",
                    "Libranza Primer Nombre",
                    "Desprendible Nomina Primer Nombre",
                    "Cedula Primer Nombre",
                    "Amortizacion Primer Nombre"
                ],
                "nombre_unificado": "Primer Nombre Unificado"
            }
        ]

        filas_correctas = []
        filas_fallos = []
        columnas_con_error = set()

        for _, row in df.iterrows():
            fila_correcta = {}
            fila_fallo = row.copy()
            errores = []
            columnas_error_fila = []

            for bloque in bloques:
                valor_unificado, cols_error = self._validar_bloque(row, bloque["columnas"], bloque["nombre_unificado"])
                if valor_unificado is not None:
                    fila_correcta[bloque["nombre_unificado"]] = valor_unificado
                else:
                    errores.append(f"No coinciden valores en {bloque['columnas']}")
                    columnas_error_fila.extend(cols_error)

            if not errores:
                for col in df.columns:
                    if all(col not in b["columnas"] for b in bloques):
                        fila_correcta[col] = row[col]
                filas_correctas.append(fila_correcta)
            else:
                fila_fallo["Error"] = " | ".join(errores)
                filas_fallos.append(fila_fallo)
                columnas_con_error.update(columnas_error_fila)  

        df_unificado = pd.DataFrame(filas_correctas)
        df_fallos = pd.DataFrame(filas_fallos)

        self.carpeta_excel_destino.mkdir(parents=True, exist_ok=True)
        self.carpeta_excel_fallos.mkdir(parents=True, exist_ok=True)

        nuevo_nombre_ok = archivo.name.replace("_reestructurado.xlsx", "_normalizado.xlsx")
        nuevo_nombre_fallos = archivo.name.replace("_reestructurado.xlsx", "_fallos.xlsx")

        ruta_ok = self.carpeta_excel_destino / nuevo_nombre_ok
        ruta_fallos = self.carpeta_excel_fallos / nuevo_nombre_fallos

        df_unificado.to_excel(ruta_ok, index=False, engine='openpyxl')
        df_fallos.to_excel(ruta_fallos, index=False, engine='openpyxl')

        if not df_fallos.empty:
            self._pintar_errores(ruta_fallos, list(columnas_con_error))

        self.logger.info(f"Archivo normalizado guardado en: {ruta_ok}")
        self.logger.info(f"Archivo de fallos guardado en: {ruta_fallos}")
        return True