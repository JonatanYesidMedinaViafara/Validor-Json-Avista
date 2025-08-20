import pandas as pd
from pathlib import Path
import logging
import re
from difflib import SequenceMatcher
from openpyxl import load_workbook
import config

# ---------- normalizadores ----------
def _norm_num_like(v):
    if pd.isna(v): return ""
    s = str(v).strip()
    try:
        if "e" in s.lower():
            return str(int(float(s)))
    except: pass
    try:
        f = float(s.replace(",", ""))
        if f.is_integer():
            return str(int(f))
    except: pass
    return s.replace(".", "").replace(",", "").replace(" ", "")

def _norm_text(v):
    if pd.isna(v): return ""
    return " ".join(str(v).strip().upper().split())

_MESES = {"ENE":"01","FEB":"02","MAR":"03","ABR":"04","MAY":"05","JUN":"06",
          "JUL":"07","AGO":"08","SEP":"09","OCT":"10","NOV":"11","DIC":"12"}

def _norm_fecha(v):
    if pd.isna(v): return ""
    s = str(v).strip().upper().replace("-", "/")
    m = re.match(r"^(\d{1,2})/([A-Z]{3})/(\d{4})$", s)
    if m and m.group(2) in _MESES:
        return f"{m.group(1).zfill(2)}/{_MESES[m.group(2)]}/{m.group(3)}"
    m2 = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m2:
        return f"{m2.group(1).zfill(2)}/{m2.group(2).zfill(2)}/{m2.group(3)}"
    return s

def _parse_date_month(x):
    """
    Devuelve un datetime al primer día del mes (mes/año); ignora el día original.
    Acepta dd/MM/yyyy o dd/MES/yyyy.
    """
    if pd.isna(x) or str(x).strip()=="":
        return None
    up = str(x).strip().upper().replace("-", "/")
    m = re.match(r"^(\d{1,2})/([A-Z]{3})/(\d{4})$", up)
    if m and m.group(2) in _MESES:
        up = f"01/{_MESES[m.group(2)]}/{m.group(3)}"
    else:
        # asume dd/mm/yyyy -> 01/mm/yyyy
        mm = re.match(r"^\d{1,2}/(\d{1,2})/(\d{4})$", up)
        if mm:
            up = f"01/{mm.group(1).zfill(2)}/{mm.group(2)}"
    dt = pd.to_datetime(up, dayfirst=True, errors="coerce")
    return None if pd.isna(dt) else dt

def _cmp(avista_val, re_val, tipo):
    if tipo == "numero":
        return _norm_num_like(avista_val) == _norm_num_like(re_val)
    if tipo == "fecha":
        return _norm_fecha(avista_val) == _norm_fecha(re_val)
    a = _norm_text(avista_val); b = _norm_text(re_val)
    return SequenceMatcher(None, a, b).ratio() >= float(config.TOLERANCIA_TEXTO)

# --- regla especial: 3 meses (solo mes/año) ---
def _max_3_meses_antes_mes_anio(fecha_desembolso_avista, fecha_vigencia_re):
    des = _parse_date_month(fecha_desembolso_avista)
    vig = _parse_date_month(fecha_vigencia_re)
    if des is None or vig is None:
        return None
    # diferencia en meses (vig <= des y (des - vig) <= 3)
    diff_months = (des.year - vig.year)*12 + (des.month - vig.month)
    return (diff_months >= 0) and (diff_months <= 3)

def _doc_out_col(doc: str) -> str:
    return "CEDULA COMPARADA" if doc.strip().upper()=="CEDULA" else doc

# ---- export format (igual) ----
FECHA_COLS   = ["FECHA VENCIMIENTO", "FECHA DESEMBOLSO", "FECHA NACIMIENTO"]
PERCENT_COLS = ["TASA NOMINAL"]
NUM_DEC_COLS = ["SALARIO", "VALOR CUOTA", "SALDO CAPITAL", "MONTO INCIAL", "CUOTA CORRIENTE"]
NUM_INT_COLS = ["CUOTAS FALTANTES", "PLAZO INICIAL"]

def _tipar_y_formatear_excel(ruta_xlsx: Path, header_row: int = 1):
    wb = load_workbook(ruta_xlsx)
    ws = wb.active
    header = {cell.value: idx+1 for idx, cell in enumerate(ws[header_row])}
    fmt_date = "DD/MM/YYYY"; fmt_pct="0.00%"; fmt_int="#,##0"; fmt_dec="#,##0.00"
    def fmt(col_name, numfmt):
        idx = header.get(col_name)
        if not idx: return
        for r in ws.iter_rows(min_row=header_row+1, min_col=idx, max_col=idx):
            for c in r: c.number_format = numfmt
    for c in FECHA_COLS: fmt(c, fmt_date)
    for c in PERCENT_COLS: fmt(c, fmt_pct)
    for c in NUM_INT_COLS: fmt(c, fmt_int)
    for c in NUM_DEC_COLS: fmt(c, fmt_dec)
    wb.save(ruta_xlsx)

# ---- nombre completo AVISTA (para comparar) ----
def _avista_nombre_completo(row) -> str:
    partes = [
        row.get("PRIMER NOMBRE",""),
        row.get("SEGUNDO NOMBRE",""),
        row.get("PRIMER APELLIDO",""),
        row.get("SEGUNDO APELLIDO",""),
    ]
    return " ".join([str(p).strip() for p in partes if isinstance(p,str) and p.strip()]).upper()

class ComparadorAvista:
    def __init__(self, carpeta_excel_reestructurado: str, carpeta_bases_avista: str, carpeta_salida: str):
        self.carpeta_excel_reestructurado = Path(carpeta_excel_reestructurado)
        self.carpeta_bases_avista = Path(carpeta_bases_avista)
        self.carpeta_salida = Path(carpeta_salida)
        self.logger = logging.getLogger("ComparadorAvista")

    def _obtener_ultimo_reestructurado(self) -> Path | None:
        archivos = list(self.carpeta_excel_reestructurado.glob("clon_json_*_reestructurado.xlsx"))
        return max(archivos, key=lambda f: f.stat().st_mtime) if archivos else None

    def _obtener_ultima_base(self) -> Path | None:
        archivos = sorted(self.carpeta_bases_avista.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        return archivos[0] if archivos else None

    def _leer_avista(self, ruta: Path):
        try:
            return pd.read_excel(ruta)
        except Exception as e:
            self.logger.exception(f"Error leyendo Avista {ruta}: {e}")
            return None

    def _leer_reestructurado(self, ruta: Path):
        try:
            return pd.read_excel(ruta, dtype=str)
        except Exception as e:
            self.logger.exception(f"Error leyendo reestructurado {ruta}: {e}")
            return None

    def comparar(self) -> bool:
        ruta_reestr = self._obtener_ultimo_reestructurado()
        if not ruta_reestr:
            self.logger.error("No hay reestructurado.")
            return False

        ruta_avista = self._obtener_ultima_base()
        if not ruta_avista:
            self.logger.error("No hay bases Avista en la carpeta configurada.")
            return False

        df_avista = self._leer_avista(ruta_avista)
        if df_avista is None or "OPERACIÓN" not in df_avista.columns:
            self.logger.error("La base Avista no contiene 'OPERACIÓN'.")
            return False

        df_res = self._leer_reestructurado(ruta_reestr)
        if df_res is None or "Numero credito" not in df_res.columns:
            self.logger.error("El reestructurado no contiene 'Numero credito'.")
            return False

        df_res["_NUM_CRED_NORM_"] = df_res["Numero credito"].apply(_norm_num_like)

        hoja = df_avista.copy(deep=True)
        out_cols = {doc: _doc_out_col(doc) for doc in config.DOCUMENTOS}
        for _, out_name in out_cols.items():
            if out_name not in hoja.columns:
                hoja[out_name] = ""

        ok_global, fallos_global = [], []

        for idx, fav in df_avista.iterrows():
            op_norm = _norm_num_like(fav.get("OPERACIÓN",""))
            cand = df_res[df_res["_NUM_CRED_NORM_"] == op_norm]
            if cand.empty:
                for out_name in out_cols.values():
                    hoja.at[idx, out_name] = "NO ENCONTRADO EN REESTRUCTURADO"
                fallos_global.append({"OPERACIÓN": fav.get("OPERACIÓN",""), "Motivo": "Sin match"})
                continue

            fila_res = cand.iloc[0]
            nombre_avista = _avista_nombre_completo(fav)
            hubo_fallos = False

            for doc, out_name in out_cols.items():
                evidencias = []
                for campo_avista, spec in config.DOCUMENTOS_MAPEO.get(doc, {}).items():
                    re_col = spec.get("re")
                    t = spec.get("tipo","texto")
                    special = spec.get("validacion_especial")
                    re_vs_re = spec.get("comparar_recontra_re", False)

                    if re_vs_re:
                        v1 = fila_res.get(spec.get("re"), "")
                        v2 = fila_res.get(spec.get("re2"), "")
                        ok = (_norm_num_like(v1) == _norm_num_like(v2))
                        evidencias.append("OK" if ok else "Son Diferente entre Solicitud Credito Solicitud y Amortizacion Numero Solicitud")
                        if not ok: hubo_fallos = True
                        continue

                    if campo_avista == "NOMBRE COMPLETO":
                        av_val = nombre_avista
                    else:
                        av_val = fav.get(campo_avista, "")

                    if not re_col or re_col not in fila_res.index:
                        evidencias.append(f"SIN DATO {campo_avista}")
                        hubo_fallos = True
                        continue

                    re_val = fila_res.get(re_col, "")

                    if special == "max_3_meses_antes_mes_anio":
                        ok = _max_3_meses_antes_mes_anio(av_val, re_val)
                        if ok is None:
                            evidencias.append(f"SIN DATO {campo_avista}")
                            hubo_fallos = True
                        elif ok:
                            evidencias.append("OK")
                        else:
                            evidencias.append("FECHA EXCEDE EL TIEMPO")
                            hubo_fallos = True
                    else:
                        ok = _cmp(av_val, re_val, t)
                        if ok:
                            evidencias.append("OK")
                        else:
                            evidencias.append("Nombre completo no coincide" if campo_avista=="NOMBRE COMPLETO"
                                              else f"FALLO {campo_avista}")
                            hubo_fallos = True

                hoja.at[idx, out_name] = ", ".join(evidencias) if evidencias else ""

            if hubo_fallos:
                fallos_global.append({"OPERACIÓN": fav.get("OPERACIÓN","")})
            else:
                ok_global.append({"OPERACIÓN": fav.get("OPERACIÓN","")})

        self.carpeta_salida.mkdir(parents=True, exist_ok=True)
        base = Path(ruta_reestr).stem.replace("_reestructurado", "")
        ruta_evid = self.carpeta_salida / f"{base}_evidencia_avista_unica.xlsx"
        ruta_ok   = self.carpeta_salida / f"{base}_ok.xlsx"
        ruta_fail = self.carpeta_salida / f"{base}_fallos.xlsx"

        hoja.to_excel(ruta_evid, index=False, engine="openpyxl")
        _tipar_y_formatear_excel(ruta_evid)
        pd.DataFrame(ok_global).to_excel(ruta_ok, index=False, engine="openpyxl")
        pd.DataFrame(fallos_global).to_excel(ruta_fail, index=False, engine="openpyxl")

        self.logger.info(f"Evidencia -> {ruta_evid}")
        self.logger.info(f"OK        -> {ruta_ok}")
        self.logger.info(f"Fallos    -> {ruta_fail}")
        return True
